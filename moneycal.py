"""#注意：原始数据第一列为文章类型，第二列为作者名字，需与 author_name 匹配，原始数据里的完成率一列需手动改为数值模式，否则会报错
#绩效表格中，需要提前准备的是：选题采用数、错字、其他三项
#没有去除实习生的团队绩效，需手动去除。若有其他人有绩效，可以在表格生成后进行统计"""


#"面线"可以作为str变量来在整体做判断

from openpyxl import load_workbook
import datetime

test = 1     #0为非调试模式，1为调试模式

date = datetime.datetime.now()
filename = str(date.year) + '-' + str(date.month - 1) + '.xlsx'
data = load_workbook(filename)
p_data = data.get_sheet_by_name('Sheet1')
p_row = p_data.max_row
#print(title_row)
author_5 = 0 #作者有几篇那几绩效，过 5 则有 1000 奖励
todaybest_quantity = 0 #今日最佳数量
todaybest_350 = 0 #今日最佳点赞超过 350 的数量
todaybest_100k = 0 #今日最佳阅读量超过 100k
p_data.cell(row = 1, column = 83).value = "影响力"

author_name = ['桑田','小花','亚南','大卷','小发','小鑫鑫','陈哲','面线','焕妍','基昊','文政','胖虎','米罗','浩然'] #根据绩效表格作相应调整
filename = '绩效.xlsx'
jixiao = load_workbook(filename)
jixiao_data = jixiao.get_sheet_by_name('Sheet1')

#头次平均阅读量初始化
first_count = 0
first_amount = 0
other_count = 0
other_amount = 0




"""头条：
影响力30w奖励600元，影响力60w奖励3000元，影响力100w奖励5000元，影响力150w奖励10000元。
好看达到2500，奖励600元，好看达到5000，奖励1200元。

次条：
二条
影响力12w奖励300元，影响力15w奖励600元，影响力20w奖励1200元。
好看达到1000，奖励400元，好看达到2000，奖励800元。

三、四条
影响力8w奖励300元，影响力12w奖励600元。
好看达到500，奖励300元，好看达到1000，奖励600元。"""
def per_article( sequence,influence,likes ): #文章序号，影响力，好看数
    performance = 0
    performance_seq = 0 #对应绩效.xlsx中的过影响力的列
    performance_like = 0  #对应绩效.xlsx中的过点赞的列
    sequence = int(sequence)
    if sequence == 1:
        if influence >= 1500000:
            performance += 10000
            performance_seq = 8
        elif influence >= 1000000:
            performance += 5000
            performance_seq = 7
        elif influence >= 600000:
            performance += 3000
            performance_seq = 6
        elif influence >= 300000:
            performance += 600
            performance_seq = 5
        else:
            performance += 0

        if likes >= 5000:
            performance += 1200
            performance_like = 4
        elif likes >= 2500:
            performance += 600
            performance_like = 3
        else:
            performance += 0

    elif sequence == 2:
        if influence >= 200000:
            performance += 1200
            performance_seq = 16
        elif influence >= 150000:
            performance += 600
            performance_seq = 15
        elif influence >= 120000:
            performance += 300
            performance_seq = 14
        else:
            performance += 0
        if likes >= 2000:
            performance += 800
            performance_like = 11
        elif likes >= 1000:
            performance += 400
            performance_like = 10
        else:
            performance += 0

    else:
        if influence >= 120000:
            performance += 600
            performance_seq = 18
        elif influence >= 80000:
            performance += 300
            performance_seq = 17
        else:
            performance += 0
        if likes >= 1000:
            performance += 600
            performance_like = 13
        elif likes >= 500:
            performance += 300
            performance_like = 12
        else:
            performance += 0

    if performance > 0:
        return [performance,1,sequence,performance_seq,performance_like]
    else:
        return [performance,0,sequence,performance_seq,performance_like]


def import_data(performance_para,author):
    if author not in author_name:
        print(author+"在绩效表中没有找到")
        return 1
    else:
        author_index = author_name.index(author) + 4
        if (performance_para[3] != 0):
            jixiao_data.cell(row = author_index, column = performance_para[3]).value += 1
        if performance_para[4] != 0:
            jixiao_data.cell(row = author_index, column = performance_para[4]).value += 1
        #jixiao_data.cell(row = author_index, column = 25).value += performance_para[0]
        jixiao_data.cell(row = author_index, column = 20).value += performance_para[1]
        if performance_para[2] == 1:
            jixiao_data.cell(row = author_index, column = 2).value += 1
        else:
            jixiao_data.cell(row = author_index, column = 9).value += 1
        return 0

for i in range(2,p_row+1):
	#影响力公式：阅读量*（1+完成率+转发率*20+好看率*30）/2
    reading_quantity = p_data.cell(row = i, column = 9).value   
    complete_rate = p_data.cell(row = i, column = 42).value
    share_rate = p_data.cell(row = i, column = 28).value
    likes_rate = p_data.cell(row = i, column = 30).value
    influence = reading_quantity*(1+complete_rate+20*share_rate+30*likes_rate)
    p_data.cell(row = i, column = 83).value = round(influence/2)

    if (p_data.cell(row = i, column = 1).value == "文章"):  
        #计算团队阅读量，团队奖励,次条阅读量 12w+ 算 12w，头条阅读量 30w+ 算 30w
        if p_data.cell(row = i, column = 4).value == '1':
            first_count += 1
            if p_data.cell(row = i, column = 9).value >= 300000:
                first_amount += 300000
            else:
                first_amount += p_data.cell(row = i, column = 9).value
        else:
            other_count += 1
            if p_data.cell(row = i, column = 9).value >= 120000:
                other_amount += 120000
            else:
                other_amount += p_data.cell(row = i, column = 9).value
        performance_article = per_article(p_data.cell(row = i, column = 4).value,p_data.cell(row = i, column = 83).value,p_data.cell(row = i, column = 25).value)
        
        if performance_article[0]>3000:
            print("超标文章")
            print(p_data.cell(row = i, column = 3).value)
            print(performance_article[0])
            print(p_data.cell(row = i, column = 83).value,p_data.cell(row = i, column = 25).value)
        import_data(performance_article,p_data.cell(row = i, column = 2).value)

    """今日最佳：
    1. 70% 的好看超过 350，桑田 500
	 2. 100%阅读量超过10万，面线 500    """
    if (p_data.cell(row = i, column = 1).value == "今日最佳"):
        todaybest_quantity += 1
        if p_data.cell(row = i, column = 25).value >= 350:
            todaybest_350 += 1
        if p_data.cell(row = i, column = 9).value >= 100000:
            todaybest_100k += 1
    
    """聊一聊：面线绩效
	1. 阅读量超过7.5万，100
	2. 留言量超过700，100 """
    if (p_data.cell(row = i, column = 1).value == "聊一聊"):
        if p_data.cell(row = i, column = 9).value >= 75000:
            if p_data.cell(row = i, column = 2).value == '聊一聊':
                if '面线' in author_name:
                    jixiao_data.cell(row = author_name.index('面线')+4, column = 21).value += 1
                    jixiao_data.cell(row = author_name.index('面线')+4, column = 25).value += 100
                else:
                    print("面线没在绩效名单中")
            else: #如果是周末聊一聊，记到对应作者头上
                if p_data.cell(row = i, column = 2).value not in author_name:
                    print (str(p_data.cell(row = i, column = 2).value) + "聊一聊没有作者名字")
                else:
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 21).value += 1
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 25).value += 100
        if p_data.cell(row = i, column = 26).value >= 700:
            if p_data.cell(row = i, column = 2).value == '聊一聊':
                if '面线' in author_name:
                    jixiao_data.cell(row = author_name.index('面线')+4, column = 21).value += 1
                    jixiao_data.cell(row = author_name.index('面线')+4, column = 25).value += 100
                else:
                    print("面线没在绩效名单中")
            else: #如果是周末聊一聊，记到对应作者头上
                if p_data.cell(row = i, column = 2).value not in author_name:
                    print (str(p_data.cell(row = i, column = 2).value) + "聊一聊没有作者名字")
                else:
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 21).value += 1
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 25).value += 100

    """    大新闻：
	1. 尾条：阅读量超过10万，100
	2. 头条：阅读量超过30万，100 """
    if (p_data.cell(row = i, column = 1).value == "大新闻"):
        if p_data.cell(row = i, column = 4).value == 1:
            if p_data.cell(row = i, column = 9).value > 300000:
                if p_data.cell(row = i, column = 2).value not in author_name:
                    print(str(p_data.cell(row = i, column = 3).value)+"这篇大新闻作者需要标注，绩效错误")
                else:
                    print(jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 19).value)
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 19).value += 1
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 25).value += 100
        else:
            if p_data.cell(row = i, column = 9).value > 100000:                
                if p_data.cell(row = i, column = 2).value not in author_name:
                    print(str(p_data.cell(row = i, column = 3).value)+"这篇文章作者需要标注，绩效错误")
                else:
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 19).value += 1
                    jixiao_data.cell(row = author_name.index(p_data.cell(row = i, column = 2).value)+4, column = 25).value += 100

#今日最佳绩效

if todaybest_100k == todaybest_quantity:
    jixiao_data.cell(row = author_name.index('面线')+4, column = 24).value += 500
    jixiao_data.cell(row = author_name.index('面线')+4, column = 26).value = str(jixiao_data.cell(row = author_name.index('面线')+4, column = 26).value)
    jixiao_data.cell(row = author_name.index('面线')+4, column = 26).value += "今日最佳100%过10万"
if todaybest_350/todaybest_quantity > 0.7:
    jixiao_data.cell(row = author_name.index('桑田')+4, column = 24).value += 500
    jixiao_data.cell(row = author_name.index('桑田')+4, column = 26).value = str(jixiao_data.cell(row = author_name.index('桑田')+4, column = 26).value)
    jixiao_data.cell(row = author_name.index('桑田')+4, column = 26).value += "今日最佳70%好看过350"


#记录头条、次条平均阅读量，头条平均达到25w，次条平均达到7.5w；同时达到以上要求，全员奖励1000元。

jixiao_data.cell(row = len(author_name)+4, column = 2).value = round(first_amount/first_count)
if first_amount/first_count > 250000:
    jixiao_data.cell(row = len(author_name)+4, column = 3).value = "达标"
else:
    jixiao_data.cell(row = len(author_name)+4, column = 3).value = "不达标"

jixiao_data.cell(row = len(author_name)+5, column = 2).value = round(other_amount/other_count)
if other_amount/other_count > 75000:
    jixiao_data.cell(row = len(author_name)+5, column = 3).value = "达标"
else:
    jixiao_data.cell(row = len(author_name)+5, column = 3).value = "不达标"
if (first_amount/first_count > 250000) & (other_amount/other_count > 75000):
    team_goal = 1000
else:
    team_goal = 0
    


#叠加错别字、团队绩效、和其他绩效
for i in range(4,3+len(author_name)):
    if jixiao_data.cell( row = i, column = 20).value >= 5:
        jixiao_data.cell( row = i, column = 20).value = 1
    else:
        jixiao_data.cell( row = i, column = 20).value = 0
    jixiao_data.cell( row = i, column = 25).value = (jixiao_data.cell( row = i, column = 25).value + jixiao_data.cell( row = i, column = 24).value + jixiao_data.cell( row = i, column = 23).value + team_goal + jixiao_data.cell( row = i, column = 22).value*50)
    toutiao_jixiao = jixiao_data.cell( row = i, column = 3).value*600 + jixiao_data.cell( row = i, column = 4).value*1200 + jixiao_data.cell( row = i, column = 5).value*600 + jixiao_data.cell( row = i, column = 6).value*3000 + jixiao_data.cell( row = i, column = 7).value*5000 + jixiao_data.cell( row = i, column = 8).value*1000
    citiao_jixiao = jixiao_data.cell( row = i, column = 10).value*400 + jixiao_data.cell( row = i, column = 11).value*800 + jixiao_data.cell( row = i, column = 12).value*300 + jixiao_data.cell( row = i, column = 13).value*600 + jixiao_data.cell( row = i, column = 14).value*300 + jixiao_data.cell( row = i, column = 15).value*600 + jixiao_data.cell( row = i, column = 16).value*1200 + jixiao_data.cell( row = i, column = 17).value*300 + jixiao_data.cell( row = i, column = 18).value*600
    jixiao_data.cell( row = i, column = 25).value = jixiao_data.cell( row = i, column = 25).value + toutiao_jixiao + citiao_jixiao
    if jixiao_data.cell( row = i, column = 20).value == 1:  #过5篇有绩效，加1000
    	jixiao_data.cell( row = i, column = 25).value += 1000
#保存绩效

jixiao.save(filename = str(date.year) + '-' + str(date.month - 1) + '绩效分析.xlsx')



	